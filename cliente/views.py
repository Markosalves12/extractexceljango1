from django.shortcuts import render
from compras.models import compras
import openpyxl
from cliente.utils import colect_dados
from django.http import HttpResponse

# Create your views here.
def view_data(request):
    objects = compras.objects.all()

    return render(
        request,
        'relatorio.html',
        {
            'objects': objects
        }
    )

def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = [
        'nome_cliente', 'sobrenome_clinte', 'valor_compra', 'data_compra'
    ]

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_title

    dados = colect_dados(request)

    for row_num, row in enumerate(dados, start=2):
        row_data = [
            row.nome_cliente,
            row.sobrenome_clinte,
            row.valor_compra,
            row.data_compra
        ]

        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(
                row=row_num,
                column=col_num
            )

            cell.value = value

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['content-disposition'] = 'attachment; filename="relatorio teste.xlsx"'

    wb.save(response)

    return response